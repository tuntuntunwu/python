import openpyxl as xl
from openpyxl.drawing.image import Image
import os

filename = "test.xlsx"
if not os.path.exists(filename):
    book = xl.Workbook()  # create a workbook
    sheet1 = book.active  # a workbook is always created with at least one worksheet. You can get it by this way
    sheet1.title = "sheet1"
    sheet2 = book.create_sheet("sheet2")  # insert at the end (default)
    book.create_sheet("sheet0", 0)  # insert at first position
    book.save(filename)  # the new workbook should be saved

book = xl.load_workbook(filename)  # load a workbook
sheet = book['sheet0']

# accessing one cell
sheet['A4'] = 4
sheet.cell(row=4, column=2, value=10)
print(sheet['A4'].value)
print(sheet['a4'].value)
print(sheet['C1'].value)
print(sheet.cell(row=4, column=2).value)

# accessing many cells
for x in range(1, 4):
    for y in range(1, 3):
        sheet.cell(row=x, column=y, value=x+y)
print("--- sheet['A'] ---")
print(sheet['A'])
print("--- sheet['A:D'] ---")
print(sheet['A:D'])
print("--- sheet['1'] ---")
print(sheet['1'])
print("---sheet['1:5'] ---")
print(sheet['1:5'])
print("--- values in sheet['A1:C4'] ---")
for row in sheet['A1:C4']:
    for cell in row:
        print(cell.value)
print("--- sheet.iter_rows() ---")
print(sheet.iter_rows(min_row=1, max_col=2, max_row=4))
print("--- sheet.iter_cols() ---")
print(sheet.iter_cols(min_row=1, max_col=2, max_row=4))
print("--- values in sheet.iter_cols() ---")
for col in sheet.iter_cols(min_row=1, max_col=2, max_row=4):
    for cell in col:
        print(cell.value)
print("--- sheet.rows ---")
print(sheet.rows)
print("--- sheet.columns ---")
print(sheet.columns)

# inserting images
img = Image("test.jpg")
sheet.add_image(img, 'A5')
# cell_w = sheet.column_dimensions['A'].width
# print(cell_w)
# cell_h = sheet.row_dimensions[1].height  # 获取单元格的宽和高
# print(cell_h)
# for i in sheet.column_dimensions['A']:
#     print(i)
# for i in sheet.row_dimensions["0"]:
#     print(i)
# print(sheet.row_dimensions[2])

book.save(filename)  # the modified workbook should be saved

